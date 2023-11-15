import logging

from hashcathelper.args import subcommand, argument, parse_config
from hashcathelper.bloodhound import CYPHER_QUERIES

log = logging.getLogger(__name__)
args = []

args.append(argument(
    '-H', '--hashes',
    default=None,
    help="path to a file containing hashes. Format: "
         "'<UPN suffix>\\<account name>:<id>:<lm hash>:<nt hash>:::'",
))

args.append(argument(
    '-A', '--accounts-plus-passwords',
    default=None,
    help="path to a file with the results from the `ntlm` subcommand. "
         "Format: '<account name>:<password>'",
))

args.append(argument(
    '-P', '--passwords-only',
    default=None,
    help="path to a file with only passwords; one per line",
))

args.append(argument(
    '-F', '--filter-accounts',
    default=None,
    help="""
path to a file containing names of accounts which are subject to analysis,
all other accounts will be filtered out (Example: only active accounts,
only kerberoastable accounts, etc.). If empty, all accounts will be
subject to analysis. Format: one per line, without domain or UPN suffix,
case insensitive.
"""
))

args.append(argument(
    '--include-disabled',
    default=False,
    action='store_true',
    help="don't exclude disabled accounts",
))

args.append(argument(
    '--include-computer-accounts',
    default=False,
    action='store_true',
    help="don't exclude computer accounts",
))

args.append(argument(
    '-B', '--bloodhound-filter',
    default=None,
    help="""
cypher query to match accounts which are subject to analysis. Requires
`--bloodhound-url` to be set. If your BloodHound database contains multiple
domains, you should set `--bloodhound-domain` as well. This parameter can
take the following values to execute a pre-defined cypher query (enabled
accounts only): """ + ', '.join(CYPHER_QUERIES.keys())
))

args.append(argument(
    '-D', '--bloodhound-domain',
    default=None,
    help="""specify the domain in BloodHound related actions
(`--bloodhound-filter`)"""
))

args.append(argument(
    '-U', '--bloodhound-url',
    default=None,
    help="""
URL to a Neo4j database containing BloodHound data. Format:
bolt[s]://<user>:<password>@<host>[:<port>]"""
))


args.append(argument(
    '-f', '--format',
    choices=['text', 'json', 'html', 'xlsx'],
    default='text',
    help="output format (default: %(default)s)",
))

args.append(argument(
    '-o', '--outfile',
    default=None,
    help="path to an output file (default: stdout)",
))

args.append(argument(
    '-m', '--pw-min-length',
    default=6,
    type=int,
    help="set minimum password length to identify accounts "
         "with 'short' passwords (default: %(default)s)",
))


args.append(argument(
    '-d', '--degree-of-detail',
    default=2,
    type=int,
    help="change the degree of detail of the report (default: %(default)s)",
))


@subcommand(args)
def analytics(args):
    '''Output interesting statistics'''
    from hashcathelper.analytics import create_report, load_lines
    from hashcathelper.bloodhound import get_driver, query_neo4j

    if args.filter_accounts:
        filter_accounts = load_lines(args.filter_accounts)
    else:
        filter_accounts = []

    if args.bloodhound_filter:
        driver = get_driver(args.bloodhound_url)
        bh_users = query_neo4j(
            driver,
            args.bloodhound_filter,
            domain=args.bloodhound_domain,
        )
    else:
        bh_users = []

    config = parse_config(args.config)
    report = create_report(
        args.hashes,
        args.accounts_plus_passwords,
        args.passwords_only,
        filter_accounts + bh_users,
        degree_of_detail=args.degree_of_detail,
        pw_min_length=args.pw_min_length,
        include_disabled=args.include_disabled,
        include_computer_accounts=args.include_computer_accounts,
        hibp_db=config.hibp_db,
    )

    if not report:
        exit(1)

    if args.format == 'xlsx':
        # xlsx is a bit special because it only contains the details
        xlsx_sanity_check(args)
        save_to_xlsx(report, args.outfile)
    else:
        out = report.export(args.format)

        if args.outfile:
            with open(args.outfile, 'w', errors='replace') as f:
                f.write(out)
        else:
            print(out, end='')


def xlsx_sanity_check(args):
    # Do some sanity checks here

    # openpyxl requires py3.6
    import sys
    if sys.version_info < (3, 6, 0):
        log.critical("XLSX format requires Python 3.6 or higher")
        exit(1)

    # Stdout does not make sense for xlsx as it's binary
    if not args.outfile:
        log.critical("XLSX format requires OUTFILE to be specified.")
        exit(1)

    # The xlsx will only contain the details, so not having the details
    # makes no sense
    if not args.degree_of_detail > 2:
        log.critical(
            "XLSX format requires degree of detail greater than 2."
        )
        exit(1)


def save_to_xlsx(report, path):
    """Saves 'details' from the report to a spreadsheet"""
    from collections import OrderedDict
    import openpyxl as pyxl
    from hashcathelper.consts import labels

    workbook = pyxl.Workbook()
    data = report._elements['details']._elements

    offset = 3

    for k, v in data.items():
        ws = workbook.create_sheet(k)
        cell = ws.cell(1, 1, labels.get(k, k))
        cell.font = pyxl.styles.Font(bold=True, size=14)

        if isinstance(v, (list, tuple)):
            for i, each in enumerate(v):
                ws.cell(offset+i+1, 1, each)
        elif isinstance(v, (dict, OrderedDict)):
            for i, key in enumerate(v.keys()):
                cell = ws.cell(offset+1, i+1, key)
                cell.font = pyxl.styles.Font(bold=True)
            for i, row in enumerate(v.values()):
                if isinstance(row, (list, tuple)):
                    for j, c in enumerate(row):
                        ws.cell(offset+j+2, i+1, c)
                else:
                    ws.cell(offset+i+1, 1, c)

    workbook.remove(workbook['Sheet'])
    workbook.save(path)
